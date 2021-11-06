import openpyxl
import math
path = "C:\\Users\\manoj\\Desktop\\Trash\\NRR.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

max_col = sheet_obj.max_column
m_row = sheet_obj.max_row

#runs scored
total_runs_scored = 0
for i in range(2,m_row+1):
    cell_obj = sheet_obj.cell(row = i, column = 2)
    # print(cell_obj.value)
    total_runs_scored = total_runs_scored + cell_obj.value

#overs played
overs_played = 0
for i in range(2,m_row+1):
    cell_obj_w = sheet_obj.cell(row=i, column=4)
    cell_obj = sheet_obj.cell(row=i, column=3)
    if cell_obj_w.value == 10:
        cell_obj.value = 20
        overs_played += cell_obj.value
    else:
        if type(cell_obj.value) is float:
            round_ = math.floor(cell_obj.value)
            decimal_ = (cell_obj.value - round_) * 10
            value_ = decimal_ / 6
            overs_played += round_ + value_
        else:
            overs_played += cell_obj.value
# print(overs_played)

#runs conceded
runs_conceded = 0
for i in range(2,m_row+1):
    cell_obj = sheet_obj.cell(row = i, column = 5)
    runs_conceded += cell_obj.value

# overs bowled
overs_bowled = 0
for i in range(2,m_row+1):
    cell_obj_w1 = sheet_obj.cell(row=i, column=6)
    cell_obj1 = sheet_obj.cell(row=i, column=7)
    if cell_obj_w1.value == 10:
        cell_obj1.value = 20
        overs_bowled += cell_obj1.value
    else:
        if type(cell_obj1.value) is float:
            round_1 = math.floor(cell_obj1.value)
            decimal_1 = (cell_obj1.value - round_1) * 10
            value_1 = decimal_1 / 6
            overs_bowled += round_1 + value_1
        else:
            overs_bowled += cell_obj1.value
# print(overs_bowled)

net_run_rate = (total_runs_scored / overs_played) - (runs_conceded / overs_bowled)

print(round(net_run_rate,3))














